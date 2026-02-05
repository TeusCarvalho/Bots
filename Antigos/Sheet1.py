# check_sheet1_first.py
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

# =========================
# CONFIG
# =========================

DEFAULT_PATH = r"C:\Users\J&T-099\OneDrive - Speed Rabbit Express Ltda\QUALIDADE_ FILIAL GO - BASE DE DADOS\04. QUALIDADE\12. BASE DE DADOS - SEM MOV NOVO"

# Nomes aceitos (case-insensitive). Se quiser aceitar também "Planilha1", adicione aqui.
ACCEPTED_FIRST_SHEET_NAMES = {"SHEET1"}


# =========================
# MODELOS
# =========================

@dataclass
class FileResult:
    file: str
    ext: str
    first_sheet: str
    ok: bool
    status: str
    detail: str


# =========================
# HELPERS
# =========================

def is_excel_file(p: Path) -> bool:
    return p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}


def should_skip(p: Path) -> bool:
    # ignora temporários do Excel tipo "~$arquivo.xlsx"
    return p.name.startswith("~$")


def normalize_sheet_name(name: str) -> str:
    return (name or "").strip().upper()


def read_first_sheet_xlsx_xlsm(path: Path) -> Tuple[Optional[str], Optional[str]]:
    """
    Retorna (first_sheet_name, error_message).
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        return None, f"openpyxl não disponível. Instale com: pip install openpyxl | erro: {e}"

    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        sheetnames = list(wb.sheetnames or [])
        wb.close()
        if not sheetnames:
            return "", None
        return sheetnames[0], None
    except Exception as e:
        return None, f"Falha ao ler workbook (xlsx/xlsm): {e}"


def read_first_sheet_xls(path: Path) -> Tuple[Optional[str], Optional[str]]:
    """
    Retorna (first_sheet_name, error_message).
    Requer xlrd (para .xls). Se não tiver, marca como erro/unsupported.
    """
    try:
        import xlrd  # type: ignore
    except Exception as e:
        return None, f"xlrd não disponível para .xls. Instale com: pip install xlrd==1.2.0 | erro: {e}"

    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheetnames = book.sheet_names()
        finally:
            book.release_resources()
        if not sheetnames:
            return "", None
        return sheetnames[0], None
    except Exception as e:
        return None, f"Falha ao ler workbook (.xls): {e}"


def check_file(path: Path) -> FileResult:
    ext = path.suffix.lower()

    if ext in {".xlsx", ".xlsm"}:
        first, err = read_first_sheet_xlsx_xlsm(path)
    elif ext == ".xls":
        first, err = read_first_sheet_xls(path)
    else:
        first, err = None, "Extensão não suportada"

    if err is not None:
        return FileResult(
            file=str(path),
            ext=ext,
            first_sheet="" if first is None else str(first),
            ok=False,
            status="ERROR",
            detail=err,
        )

    first_sheet = "" if first is None else str(first)
    ok = normalize_sheet_name(first_sheet) in ACCEPTED_FIRST_SHEET_NAMES

    return FileResult(
        file=str(path),
        ext=ext,
        first_sheet=first_sheet,
        ok=ok,
        status="OK" if ok else "FAIL",
        detail="",
    )


def iter_files(base: Path, recursive: bool) -> List[Path]:
    if recursive:
        files = [p for p in base.rglob("*") if p.is_file()]
    else:
        files = [p for p in base.iterdir() if p.is_file()]
    files = [p for p in files if is_excel_file(p) and not should_skip(p)]
    return sorted(files)


def write_csv(out_path: Path, results: List[FileResult]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["file", "ext", "first_sheet", "ok", "status", "detail"])
        for r in results:
            w.writerow([r.file, r.ext, r.first_sheet, "1" if r.ok else "0", r.status, r.detail])


# =========================
# MAIN
# =========================

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Verifica se a primeira aba (1ª sheet) de cada Excel é 'Sheet1'. Gera relatório CSV."
    )
    ap.add_argument("--path", default=DEFAULT_PATH, help="Pasta base para varrer os arquivos")
    ap.add_argument("--recursive", action="store_true", help="Varre subpastas também")
    ap.add_argument(
        "--out",
        default="relatorio_check_sheet1.csv",
        help="Nome/caminho do CSV de saída (padrão: relatorio_check_sheet1.csv na pasta alvo)",
    )
    args = ap.parse_args()

    base = Path(args.path)
    if not base.exists() or not base.is_dir():
        print(f"[ERRO] Pasta inválida: {base}")
        return 2

    files = iter_files(base, recursive=args.recursive)
    if not files:
        print("[INFO] Nenhum arquivo Excel encontrado (.xlsx/.xlsm/.xls).")
        return 0

    results: List[FileResult] = []
    for i, p in enumerate(files, 1):
        r = check_file(p)
        results.append(r)
        print(f"[{i}/{len(files)}] {r.status} | {p.name} | first_sheet='{r.first_sheet}'" + (f" | {r.detail}" if r.detail else ""))

    ok_count = sum(1 for r in results if r.status == "OK")
    fail_count = sum(1 for r in results if r.status == "FAIL")
    err_count = sum(1 for r in results if r.status == "ERROR")

    out_path = Path(args.out)
    if not out_path.is_absolute():
        out_path = base / out_path

    write_csv(out_path, results)

    print("\n==============================")
    print(f"Total: {len(results)}")
    print(f"OK   : {ok_count}")
    print(f"FAIL : {fail_count}")
    print(f"ERROR: {err_count}")
    print(f"Relatório: {out_path}")
    print("==============================\n")

    return 0 if (fail_count == 0 and err_count == 0) else 1


if __name__ == "__main__":
    raise SystemExit(main())
