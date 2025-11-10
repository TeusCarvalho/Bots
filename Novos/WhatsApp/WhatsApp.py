# WhatsApp_v5.5_defensive.py
# ------------------------------------------------------------
# J&T Express – Bot de cadastro de motoristas (Black Friday)
# - v5.1: Finaliza contato processando emails de notificação
# - v5.2: Lê números de telefone de uma planilha Excel
# - v5.3: Sistema de hash local para controle de duplicação (offline)
# - v5.4: Camada defensiva robusta contra bloqueios + fila persistente
# - v5.5: Melhorias de performance, monitoramento e usabilidade (Rich, Polars, Watchdog, etc.)
# - v5.5.1: CORREÇÃO - Função de dependências agora trata essenciais vs. opcionais.
# - v5.5.2: CORREÇÃO FINAL - logger.success trocado por logger.info para compatibilidade.
# ------------------------------------------------------------

# --- BIBLIOTECAS ESSENCIAIS ---
import os, csv, time, random, hashlib, re, sys, json, shutil
import imaplib
import email
from email.header import decode_header
from datetime import datetime, date, timedelta
from collections import defaultdict, deque

# --- BIBLIOTECAS SELENIUM ---
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# --- BIBLIOTECAS DEFENSIVAS E DE LOG ---
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Bibliotecas para Excel não instaladas. Execute: pip install pandas openpyxl")

try:
    from loguru import logger
    logger.remove()
    logger.add(sys.stdout,
               format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <8}</level> | <cyan>{message}</cyan>",
               level="INFO")
    logger.add("logs/{time:YYYY-MM-DD}.log", rotation="1 day", level="INFO", enqueue=True, encoding="utf-8")
except ImportError:
    import logging
    logger = logging.getLogger(__name__)
    logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(message)s",
                        datefmt="%Y-%m-%d %H:%M:%S")
    print("⚠️ Biblioteca 'loguru' não instalada. Usando logging padrão. Instale com: pip install loguru")

try:
    from tenacity import retry, stop_after_attempt, wait_random
except ImportError:
    print("⚠️ Biblioteca 'tenacity' não instalada. O recurso de retry automático pode não funcionar. Instale com: pip install tenacity")

try:
    from filelock import FileLock
except ImportError:
    print("⚠️ Biblioteca 'filelock' não instalada. Execute: pip install filelock")

# --- BIBLIOTECAS ADICIONAIS v5.5 ---
try:
    import psutil
except ImportError:
    print("⚠️ Biblioteca 'psutil' não instalada. A verificação do processo Chrome será desativada. Instale com: pip install psutil")
    psutil = None

try:
    from rich.console import Console
    from rich.table import Table
    RICH_AVAILABLE = True
except ImportError:
    print("⚠️ Biblioteca 'rich' não instalada. A tabela inicial não será exibida. Instale com: pip install rich")
    RICH_AVAILABLE = False

try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False


# =========================
# FUNÇÕES DE AMBIENTE E CONFIGURAÇÃO (v5.5.2 - CORRIGIDA)
# =========================
def checar_dependencias():
    """
    Verifica se as dependências essenciais estão instaladas.
    Bibliotecas opcionais apenas emitem um aviso se faltantes.
    """
    logger.info("🔍 Verificando dependências...")

    # Dependências essenciais (o script não funciona sem elas)
    essenciais = {
        "selenium": "selenium", "pandas": "pandas", "openpyxl": "openpyxl",
        "tenacity": "tenacity", "filelock": "filelock", "webdriver_manager": "webdriver-manager"
    }
    faltantes_essenciais = []
    for nome_modulo, nome_pacote in essenciais.items():
        try:
            __import__(nome_modulo)
        except ImportError:
            faltantes_essenciais.append(nome_pacote)

    if faltantes_essenciais:
        logger.error(f"❌ Dependências ESSENCIAIS faltantes: {', '.join(faltantes_essenciais)}")
        logger.error("Por favor, instale-as com: pip install " + " ".join(faltantes_essenciais))
        sys.exit(1)

    # Dependências opcionais (apenas melhoram a experiência)
    opcionais = {
        "loguru": "loguru", "rich": "rich", "polars": "polars", "psutil": "psutil"
    }
    faltantes_opcionais = []
    for nome_modulo, nome_pacote in opcionais.items():
        try:
            __import__(nome_modulo)
        except ImportError:
            faltantes_opcionais.append(nome_pacote)

    if faltantes_opcionais:
        logger.warning(f"⚠️ Dependências OPCIONAIS faltantes: {', '.join(faltantes_opcionais)}")
        logger.warning("O bot funcionará sem elas, mas recursos avançados (logs coloridos, tabelas, performance) serão desativados.")
        logger.warning(f"Para uma experiência completa, instale com: pip install " + " ".join(faltantes_opcionais))

    # v5.5.2: CORREÇÃO - Usando logger.info() para compatibilidade com logging padrão
    logger.info("✅ Verificação de dependências concluída.")


def exibir_tabela_inicial():
    """Exibe uma tabela com as configurações das sessões usando Rich."""
    if not RICH_AVAILABLE:
        logger.info("Biblioteca 'rich' não disponível para exibir a tabela.")
        return

    console = Console()
    table = Table(title="🚀 WhatsApp v5.5 Defensive - Sessões Ativas", show_header=True, header_style="bold magenta")
    table.add_column("Sessão", style="cyan", no_wrap=True)
    table.add_column("Perfil", style="magenta")
    table.add_column("Limite", justify="right")

    for s in SESSOES:
        table.add_row(s['nome'], s['user_data_dir'], str(s['limite']))

    console.print(table)


# =========================
# CONFIGURAÇÕES PRINCIPAIS
# =========================
SESSOES = [
    {"nome": "Conta_1", "user_data_dir": "C:/WhatsAppSession_1", "limite": 175},
    {"nome": "Conta_2", "user_data_dir": "C:/WhatsAppSession_2", "limite": 175},
]

# ======================================================
# 📥 LEITURA DE PLANILHA DE TELEFONES (v5.2)
# ======================================================
PLANILHA_NUMEROS = r"C:\Users\J&T-099\OneDrive - Speed Rabbit Express Ltda (1)\Área de Trabalho\Testes\Telefones\telefones_motoristas.xlsx"
UF_PADRAO = "DF"

# --- CONFIGURAÇÕES DE COMPORTAMENTO E DEFESA ---
HEADLESS = False
REENVIO_ATIVO = True
HORA_INICIO, HORA_FIM = 8, 18
FINALIZAR_AO_TERMINAR = True
INATIVIDADE_AUTOEXIT_MIN = 120

# --- CONFIGURAÇÕES DE AUTOMAÇÃO DIÁRIA ---
REINICIAR_DIARIAMENTE = True
MODO_PLANTAO = True
LIMPEZA_SEMANAL_ATIVA = True
DIA_LIMPEZA_SEMANAL = 6
HORA_LIMPEZA_SEMANAL = 23

# --- CONFIGURAÇÕES DE LOGS E RELATÓRIOS ---
SALVAR_EM_EXCEL = True
EXCEL_FILE = "Base_Master_Motoristas.xlsx"
GERAR_RELATORIO_DIARIO = True
GERAR_RELATORIO_SEMANAL = True

# --- CONFIGURAÇÕES DE NOTIFICAÇÕES ---
NOTIFICACAO_ATIVA = False
WEBHOOK_URL = ""

# --- CONFIGURAÇÕES DE FINALIZAÇÃO (via Email) ---
FINALIZAR_COM_FORMULARIO = True
INTERVALO_CHECAGEM_EMAIL_SEG = 300
MSG_FORMULARIO_RECEBIDO = "🎉 Ótimo! Recebemos seu cadastro. Em breve, nossa equipe entrará em contato com os próximos passos. Muito obrigado!"
HASH_FILE = "processed_hashes.json"

# --- CONFIGURAÇÕES DE EMAIL (PREENCHA COM SEUS DADOS) ---
EMAIL_IMAP_SERVER = "imap.gmail.com"
EMAIL_IMAP_PORT = 993
EMAIL_ADDRESS = ""
EMAIL_PASSWORD = ""
EMAIL_FOLDER = "INBOX"
EMAIL_SUBJECT_FILTER = "Google Forms"

# --- PASTAS E ARQUIVOS DE DADOS ---
PASTA_PRINTS = "prints_motoristas"
LOG_CSV = "log_motoristas.csv"
DNC_FILE = "nao_enviar_mais.csv"
AGENDA_REENVIO = "agenda_reenvio.csv"
STATE_FILE = "state.json"
STATE_LOCK = "state.json.lock"
os.makedirs(PASTA_PRINTS, exist_ok=True)
os.makedirs("logs", exist_ok=True)

# --- v5.4: CONFIGURAÇÕES DEFENSIVAS ---
INTERVALO_CURTO = (35, 70)
INTERVALO_MEDIO = (90, 180)
INTERVALO_LONGO = (300, 600)
MICROPAUSA_TECLA = (0.05, 0.15)

# --- v5.5: CONFIGURAÇÕES ADICIONAIS ---
MODO_TESTE = False  # Define como True para simular envios sem enviar
DIAS_LIMPEZA_PRINTS = 3  # Limpa prints com mais de 3 dias
UFS_PRIORITARIAS = ["SP", "RJ", "MG", "DF"]  # Fila inteligente por prioridade

# --- MENSAGENS ---
MSG_INICIAL_VARIACOES = [
    "Olá! Somos a J&T Express 🚚. Está disponível para novas coletas?",
    "Bom dia! 👋 Aqui é da J&T Express. Estamos cadastrando motoristas parceiros.",
    "Oi, tudo bem? 😊 A J&T Express está com bônus especiais na Black Friday!",
    "Fala parceiro! 🚛 Temos novas rotas disponíveis. Quer saber mais?",
    "Olá! 📦 A J&T Express está com oportunidades de entrega nesta semana."
]
MSG_INTERESSADO = "Perfeito! 🙌 Nosso time entrará em contato em breve. Por favor, preencha o formulário abaixo para acelerar seu cadastro:"
MSG_NAO_INTERESSADO = "Sem problema 👍 Agradecemos o seu tempo. Boa sorte e sucesso na estrada!"
MSG_PEDIR_CONF = "Por favor, responda apenas com 'sim' ou 'não' 👍"
MSG_REENVIO_12H = "Oi! 👋 Relembrando a campanha Black Friday: bônus de até R$ 1,30 + R$ 0,50 extra por pacote. Responda 'Sim' para saber mais!"
MSG_FORA_HORARIO_NOITE = "Olá! Nosso horário são das 8h às 18h. Amanhã nossa equipe entrará em contato. 😊"
MSG_FORA_HORARIO_FDS = "Olá! Nosso horário são de segunda a sexta, das 8h às 18h. Segunda retornaremos o contato. 🚛"
LINK_FORM = "https://forms.gle/qckjgW3GkRiJ8uU56"

# --- v5.4: CONFIGURAÇÕES DO DRIVER ---
DEFAULT_WINDOW_POSITION = (1920, 0)
DEFAULT_WINDOW_SIZE = (1100, 800)
DEFAULT_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

# Mapeamento de DDD por UF
DDD_UF = {"AC": "68", "AL": "82", "AP": "96", "AM": "92", "BA": "71", "CE": "85", "DF": "61", "ES": "27", "GO": "62",
          "MA": "98", "MT": "65", "MS": "67", "MG": "31", "PA": "91", "PB": "83", "PR": "41", "PE": "81", "PI": "86",
          "RJ": "21", "RN": "84", "RS": "51", "RO": "69", "RR": "95", "SC": "48", "SP": "11", "SE": "79", "TO": "63"}


# =========================
# GERENCIADOR DE FILA PERSISTENTE (v5.5 com Polars e Prioridade)
# =========================
class QueueManager:
    def __init__(self):
        self.state = {}
        self.dnc = set()
        self.enviados_hoje = set()
        self.lock = FileLock(STATE_LOCK, timeout=5)
        self._load_state()
        self._load_dnc()
        self._load_enviados_hoje()

    def _load_state(self):
        try:
            with self.lock:
                if os.path.exists(STATE_FILE):
                    with open(STATE_FILE, "r", encoding="utf-8") as f:
                        self.state = json.load(f)
                else:
                    self.state = {}
        except Exception as e:
            logger.error(f"Erro ao carregar estado: {e}")
            self.state = {}

    def _save_state(self):
        try:
            with self.lock:
                with open(STATE_FILE, "w", encoding="utf-8") as f:
                    json.dump(self.state, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.error(f"Erro ao salvar estado: {e}")

    def _load_dnc(self):
        try:
            if os.path.exists(DNC_FILE):
                with open(DNC_FILE, encoding="utf-8") as f:
                    self.dnc = set(line.strip() for line in f if line.strip())
        except Exception as e:
            logger.error(f"Erro ao carregar DNC: {e}")
            self.dnc = set()

    def _load_enviados_hoje(self):
        try:
            hoje = datetime.now().date().isoformat()
            if os.path.exists(LOG_CSV):
                with open(LOG_CSV, encoding="utf-8") as f:
                    for r in csv.DictReader(f):
                        if r.get("timestamp", "").startswith(hoje) and "ENVIO_OK" in r.get("etapa", ""):
                            self.enviados_hoje.add(r.get("numero"))
        except Exception as e:
            logger.error(f"Erro ao carregar enviados hoje: {e}")
            self.enviados_hoje = set()

    def normalize_number(self, raw, uf_default="DF"):
        if not raw: return None
        digits = re.sub(r'\D', '', str(raw))
        if len(digits) in (8, 9):
            ddd = DDD_UF.get(uf_default.upper(), "61")
            digits = ddd + digits
        if not digits.startswith("55"): digits = "55" + digits
        if len(digits) in (12, 13): return "+" + digits
        return None

    def load_numbers_from_excel(self, path=PLANILHA_NUMEROS):
        if not os.path.exists(path):
            logger.error(f"Planilha não encontrada: {path}")
            return []

        df = None
        try:
            # v5.5: Tenta usar Polars para performance
            if POLARS_AVAILABLE:
                logger.info("📊 Usando Polars para ler a planilha (performance melhorada).")
                df_pl = pl.read_excel(path)
                df = df_pl.to_pandas()
            else:
                raise ImportError("Polars não disponível, usando Pandas.")
        except Exception:
            logger.info("📊 Usando Pandas para ler a planilha.")
            df = pd.read_excel(path, dtype=str)

        df.columns = [c.strip().lower() for c in df.columns]

        col_num = next((c for c in df.columns if "número" in c or "telefone" in c), None)
        col_uf = next((c for c in df.columns if "estado" in c or "uf" in c), None)

        rows = []
        for _, r in df.iterrows():
            raw = r.get(col_num) if col_num else None
            uf = (r.get(col_uf) if col_uf else None) or "DF"
            num = self.normalize_number(raw, uf_default=uf)
            if num:
                rows.append({"nome": r.get('nome', ''), "numero": num, "uf": uf})

        seen = set()
        out = []
        for x in rows:
            if x["numero"] not in seen:
                seen.add(x["numero"])
                out.append(x)
        return out

    def init_state_from_sheet(self):
        rows = self.load_numbers_from_excel()
        changed = False
        for r in rows:
            num = r["numero"]
            if num not in self.state:
                self.state[num] = {
                    "status": "pending", "attempts": 0, "last_attempt": None,
                    "session": None, "nome": r.get("nome", ""), "uf": r.get("uf", "")
                }
                changed = True
        if changed: self._save_state()
        return self.state

    def get_next_number_prioritario(self, session_name, max_attempts=3):
        """Busca o próximo número de UFs prioritárias."""
        for uf in UFS_PRIORITARIAS:
            for num, meta in self.state.items():
                if num in self.enviados_hoje or num in self.dnc: continue
                if meta.get("uf") == uf and meta.get("status") in ("pending", "failed") and meta.get("attempts", 0) < max_attempts:
                    self.state[num]["status"] = "in_progress"
                    self.state[num]["session"] = session_name
                    self.state[num]["last_attempt"] = datetime.utcnow().isoformat()
                    self._save_state()
                    return num, meta
        return None, None

    def get_next_number_for_session(self, session_name, max_attempts=3):
        self._load_state()
        self._load_dnc()
        self._load_enviados_hoje()
        self._reclaim_stuck_claims()

        # v5.5: Tenta pegar um número prioritário primeiro
        num, meta = self.get_next_number_prioritario(session_name, max_attempts)
        if num: return num, meta

        # Se não houver prioritários, pega qualquer um
        for num, meta in self.state.items():
            if num in self.enviados_hoje or num in self.dnc: continue
            if meta.get("status") in ("pending", "failed") and meta.get("attempts", 0) < max_attempts:
                self.state[num]["status"] = "in_progress"
                self.state[num]["session"] = session_name
                self.state[num]["last_attempt"] = datetime.utcnow().isoformat()
                self._save_state()
                return num, meta
        return None, None

    def _reclaim_stuck_claims(self, timeout_minutes=60):
        changed = False
        now = datetime.utcnow()
        for num, meta in self.state.items():
            if meta.get("status") == "in_progress" and meta.get("last_attempt"):
                try:
                    last_attempt = datetime.fromisoformat(meta["last_attempt"])
                    if (now - last_attempt).total_seconds() > timeout_minutes * 60:
                        self.state[num]["status"] = "pending"
                        self.state[num]["session"] = None
                        changed = True
                except:
                    pass
        if changed: self._save_state()

    def mark_sent(self, numero, session_name):
        self._load_state()
        if numero not in self.state: self.state[numero] = {}
        self.state[numero].update({
            "status": "sent", "attempts": self.state.get(numero, {}).get("attempts", 0) + 1,
            "last_attempt": datetime.utcnow().isoformat(), "session": session_name
        })
        self._save_state()

    def mark_failed(self, numero, session_name, reason="error"):
        self._load_state()
        if numero not in self.state: self.state[numero] = {}
        self.state[numero].update({
            "status": "failed", "attempts": self.state.get(numero, {}).get("attempts", 0) + 1,
            "last_attempt": datetime.utcnow().isoformat(), "session": session_name, "reason": reason
        })
        self._save_state()

    def mark_finalized(self, numero):
        self._load_state()
        if numero not in self.state: self.state[numero] = {}
        self.state[numero].update({"status": "finalized", "last_attempt": datetime.utcnow().isoformat()})
        self._save_state()

    def add_to_dnc(self, numero):
        if numero not in self.dnc:
            with open(DNC_FILE, "a", encoding="utf-8") as f: f.write(numero + "\n")
            self.dnc.add(numero)

    def get_statistics(self):
        stats = defaultdict(int)
        for num, meta in self.state.items():
            stats[meta.get("status", "unknown")] += 1
        return dict(stats)


queue_manager = QueueManager()

# =========================
# CONTROLE DE ESTADO E DUPLICAÇÃO (v5.5 com Hash Rastreável)
# =========================
processed_hashes = {}
hash_meta = {}


def carregar_hashes_processados():
    global processed_hashes, hash_meta
    if os.path.exists(HASH_FILE):
        try:
            with open(HASH_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                processed_hashes = set(data.get("hashes", []))
                hash_meta = data.get("meta", {})
                logger.info(f"🔑 Carregados {len(processed_hashes)} hashes processados.")
        except Exception as e:
            logger.error(f"Erro ao carregar hashes: {e}")
            processed_hashes = set()
            hash_meta = {}


def salvar_hashes_processados():
    try:
        with open(HASH_FILE, "w", encoding="utf-8") as f:
            json.dump({"hashes": list(processed_hashes), "meta": hash_meta}, f, indent=2)
    except Exception as e:
        logger.error(f"Erro ao salvar hashes: {e}")


def gerar_hash_local(numero, etapa, timestamp=None):
    if timestamp is None: timestamp = datetime.utcnow().isoformat()
    try:
        base = f"{numero}_{etapa}_{timestamp}"
        return hashlib.sha256(base.encode()).hexdigest()
    except Exception as e:
        logger.error(f"Erro ao gerar hash local: {e}")
        return None


def registrar_hash_evento(numero, etapa):
    h = gerar_hash_local(numero, etapa)
    if not h or h in processed_hashes: return False
    processed_hashes.add(h)
    hash_meta[h] = {"numero": numero, "etapa": etapa}
    salvar_hashes_processados()
    logger.info(f"🧩 Hash registrado: {numero} ({etapa})")
    return True


def ja_registrado(numero, etapa):
    # v5.5: Verificação mais robusta usando metadados
    for h, meta in hash_meta.items():
        if meta.get("numero") == numero and meta.get("etapa") == etapa:
            return True
    return False


# =========================
# FUNÇÕES DEFENSIVAS E DE COMPORTAMENTO HUMANO (v5.4)
# =========================
def escolher_pausa():
    hora = datetime.now().hour
    if 7 <= hora < 10:
        return random.uniform(*INTERVALO_CURTO)
    elif 10 <= hora < 18:
        return random.uniform(*INTERVALO_MEDIO)
    else:
        return random.uniform(*INTERVALO_LONGO)


def escolher_mensagem(): return random.choice(MSG_INICIAL_VARIACOES)


def digitar_humano(campo, texto):
    base_delay = random.uniform(*MICROPAUSA_TECLA)
    for ch in texto:
        campo.send_keys(ch)
        time.sleep(base_delay + random.uniform(0, 0.12))
    if random.random() < 0.2: time.sleep(random.uniform(0.5, 1.5))


def simular_acao_humana(driver):
    try:
        acao = random.random()
        if acao < 0.1:
            driver.execute_script("window.scrollBy(0, 300);")
        elif acao < 0.2:
            driver.execute_script("window.scrollBy(0, -300);")
        elif acao < 0.25:
            driver.refresh(); time.sleep(5)
        elif acao < 0.3:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(1, 3))
    except Exception as e:
        logger.warning(f"⚠️ Falha na simulação de ação humana: {e}")


# =========================
# DRIVER SEGURO E DISCRETO (v5.4)
# =========================
def criar_driver_discreto(sessao, profile_dir=None, proxy=None, user_agent=None,
                          window_position=DEFAULT_WINDOW_POSITION, window_size=DEFAULT_WINDOW_SIZE, headless=False):
    if not user_agent: user_agent = random.choice(DEFAULT_USER_AGENTS)
    opts = Options()
    if profile_dir is None: profile_dir = sessao.get("user_data_dir")
    if profile_dir:
        os.makedirs(profile_dir, exist_ok=True)
        opts.add_argument(f"--user-data-dir={os.path.abspath(profile_dir)}")
    opts.add_argument(f"--window-position={window_position[0]},{window_position[1]}")
    opts.add_argument(f"--window-size={window_size[0]},{window_size[1]}")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option('useAutomationExtension', False)
    opts.add_argument(f"--user-agent={user_agent}")
    if proxy: opts.add_argument(f"--proxy-server={proxy}")
    # caps = DesiredCapabilities.CHROME # Comentado, pois pode ser obsoleto em versões recentes
    # caps["goog:loggingPrefs"] = {"browser": "OFF", "performance": "OFF"}
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts) # , desired_capabilities=caps)
    except Exception as e:
        logger.error(f"Erro ao iniciar ChromeDriver: {e}")
        raise
    try:
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', { get: () => undefined }); window.navigator.chrome = { runtime: {} }; Object.defineProperty(navigator, 'languages', { get: () => ['pt-BR','pt','en-US','en'] }); Object.defineProperty(navigator, 'plugins', { get: () => [1,2,3,4,5] });"})
    except Exception:
        logger.debug("Falha ao aplicar CDP anti-detect (não crítico).")
    try:
        driver.get("https://web.whatsapp.com/")
        WebDriverWait(driver, 15).until(lambda d: d.execute_script("return document.readyState === 'complete'"))
        time.sleep(1.5)
    except Exception:
        logger.warning("A página pode demorar a carregar. Continue...")
    logger.info(f"Sessão '{sessao.get('nome')}' iniciada.")
    return driver


def sessao_ativa(driver, timeout=8):
    try:
        XPATH_INPUT = '//footer//div[@contenteditable="true" and @role="textbox"]'
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, XPATH_INPUT)))
        return True
    except Exception:
        return False


def criar_perfil_se_nao_existir(sessao):
    profile_dir = sessao.get("user_data_dir")
    if profile_dir and not os.path.exists(profile_dir):
        os.makedirs(profile_dir, exist_ok=True)
        logger.info(f"Criado perfil para {sessao['nome']}: {profile_dir}")


# v5.5: Verificação de processo Chrome
def chrome_ativo():
    if not psutil: return True  # Assume ativo se não puder verificar
    for proc in psutil.process_iter(['name']):
        if 'chrome' in proc.info['name'].lower(): return True
    return False


# v5.5: Limpeza de prints antigos
def limpar_prints_antigos(dias=DIAS_LIMPEZA_PRINTS):
    try:
        limite = datetime.now() - timedelta(days=dias)
        removidos = 0
        for f in os.listdir(PASTA_PRINTS):
            path = os.path.join(PASTA_PRINTS, f)
            if os.path.isfile(path) and os.path.getmtime(path) < limite.timestamp():
                os.remove(path)
                removidos += 1
        if removidos > 0:
            logger.info(f"🧹 Limpados {removidos} prints antigos (mais de {dias} dias).")
    except Exception as e:
        logger.error(f"Erro ao limpar prints antigos: {e}")


# =========================
# FUNÇÕES DE LOG E RELATÓRIOS
# =========================
def escrever_log(ts, numero, etapa, detalhe="", sessao=""):
    novo = not os.path.exists(LOG_CSV)
    with open(LOG_CSV, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if novo: w.writerow(["timestamp", "numero", "etapa", "detalhe", "sessao"])
        w.writerow([ts, numero, etapa, detalhe, sessao])
    if SALVAR_EM_EXCEL and EXCEL_AVAILABLE: escrever_log_excel(ts, numero, etapa, detalhe, sessao)


def escrever_log_excel(ts, numero, etapa, detalhe="", sessao=""):
    try:
        hoje = date.today().isoformat()
        if not os.path.exists(EXCEL_FILE):
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                pd.DataFrame(columns=["timestamp", "numero", "etapa", "detalhe", "sessao"]).to_excel(writer,
                                                                                                     sheet_name=hoje,
                                                                                                     index=False)
        book = load_workbook(EXCEL_FILE)
        if hoje not in book.sheetnames:
            book.create_sheet(hoje)
            sheet = book[hoje]
            sheet.append(["timestamp", "numero", "etapa", "detalhe", "sessao"])
            for cell in sheet[1]: cell.font = Font(bold=True); cell.fill = PatternFill(start_color="366092",
                                                                                       end_color="366092",
                                                                                       fill_type="solid"); cell.font = Font(
                color="FFFFFF", bold=True); cell.alignment = Alignment(horizontal="center")
        sheet = book[hoje]
        sheet.append([ts, numero, etapa, detalhe, sessao])
        book.save(EXCEL_FILE)
    except Exception as e:
        logger.error(f"Erro ao escrever no Excel: {e}")


def gerar_relatorio_diario():
    if not GERAR_RELATORIO_DIARIO: return
    hoje = date.today().isoformat()
    stats = {"envios": 0, "respostas": 0, "interessados": 0, "nao_interessados": 0, "reenvios": 0, "formularios": 0}
    if not os.path.exists(LOG_CSV): return stats
    with open(LOG_CSV, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row["timestamp"].startswith(hoje):
                etapa = row["etapa"]
                if "ENVIO_OK" in etapa:
                    stats["envios"] += 1
                elif "RESPONDEU_" in etapa:
                    stats["respostas"] += 1
                    if etapa == "RESPONDEU_SIM":
                        stats["interessados"] += 1
                    elif etapa == "RESPONDEU_NAO":
                        stats["nao_interessados"] += 1
                elif "REENVIO_12H" in etapa:
                    stats["reenvios"] += 1
                elif "FORMULARIO_PREENCHIDO" in etapa:
                    stats["formularios"] += 1
    with open(f"relatorio_{hoje}.json", "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2)
    return stats


# =========================
# FUNÇÕES AUXILIARES
# =========================
def ja_enviado(numero): return numero in queue_manager.enviados_hoje


def foi_finalizado(numero): return queue_manager.state.get(numero, {}).get("status") == "finalized"


def salvar_print(driver, tag):
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(PASTA_PRINTS, f"{tag}_{ts}.png")
        driver.save_screenshot(path)
        logger.info(f"📸 {path}")
        return path
    except Exception as e:
        logger.warning(f"Erro ao salvar print: {e}")
        return None


# =========================
# SELENIUM HELPERS + ENVIO SEGURO (v5.5 com Dry Run)
# =========================
XPATH_INPUT = '//footer//div[@contenteditable="true" and @role="textbox"]'
XPATH_MSG_IN = '//div[contains(@class,"message-in")]//span[@dir="ltr"]'


def limpar_texto_bmp(s): return ''.join(ch for ch in s if ord(ch) <= 0xFFFF)


def abrir_whatsapp(driver):
    driver.get("https://web.whatsapp.com/")
    logger.info("🔄 Aguardando login no WhatsApp Web (20 s)…")
    time.sleep(20)


def abrir_chat(driver, numero):
    driver.get(f"https://web.whatsapp.com/send?phone={numero}")
    WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.XPATH, XPATH_INPUT)))


@retry(stop=stop_after_attempt(3), wait=wait_random(min=3, max=7))
def enviar_texto_seguro(driver, texto):
    # v5.5: Modo Teste (Dry Run)
    if MODO_TESTE:
        logger.info(f"[MODO TESTE] Simulação de envio: '{texto[:50]}...'")
        return

    texto = limpar_texto_bmp(texto)
    campo = driver.find_element(By.XPATH, XPATH_INPUT)
    campo.click()
    time.sleep(0.5)
    try:
        campo.clear()
    except Exception:
        pass
    digitar_humano(campo, texto)
    time.sleep(random.uniform(0.6, 1.2))
    campo.send_keys(Keys.ENTER)
    time.sleep(1.0)
    logger.info("✅ Mensagem enviada com sucesso.")


# =========================
# AGENDA DE REENVIO
# =========================
def agendar_reenvio(numero, horas=12):
    when = (datetime.utcnow() + timedelta(hours=horas)).isoformat()
    novo = not os.path.exists(AGENDA_REENVIO)
    with open(AGENDA_REENVIO, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if novo: w.writerow(["numero", "iso_schedule"])
        w.writerow([numero, when])
    logger.info(f"🕐 Reenvio agendado para {numero} em {horas} h")


def carregar_agenda():
    if not os.path.exists(AGENDA_REENVIO): return []
    with open(AGENDA_REENVIO, encoding="utf-8") as f: return list(csv.DictReader(f))


def salvar_agenda(itens):
    with open(AGENDA_REENVIO, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["numero", "iso_schedule"])
        for it in itens: w.writerow([it["numero"], it["iso_schedule"]])


def processar_agenda_reenvio(driver):
    if not REENVIO_ATIVO: return
    itens = carregar_agenda()
    if not itens: return
    now = datetime.utcnow()
    due, keep = [], []
    for it in itens:
        try:
            when = datetime.fromisoformat(it["iso_schedule"])
            (due if when <= now else keep).append(it)
        except Exception:
            continue
    if not due: return
    for it in due:
        numero = it["numero"]
        try:
            abrir_chat(driver, numero)
            enviar_texto_seguro(driver, MSG_REENVIO_12H)
            if not ja_registrado(numero, "REENVIO_12H"):
                registrar_hash_evento(numero, "REENVIO_12H")
                escrever_log(datetime.utcnow().isoformat(), numero, "REENVIO_12H", "Lembrete enviado")
            salvar_print(driver, f"reenviou12h_{numero[-4:]}")
        except Exception as e:
            logger.error(f"Erro no reenvio 12 h para {numero}: {e}")
    salvar_agenda(keep)


# =========================
# CONTROLE DE HORÁRIO E AUTOMAÇÃO
# =========================
def dentro_do_horario():
    now = datetime.now()
    return now.weekday() < 5 and HORA_INICIO <= now.hour < HORA_FIM


def fora_horario_mensagem():
    now = datetime.now()
    return MSG_FORA_HORARIO_NOITE if now.weekday() < 5 else MSG_FORA_HORARIO_FDS


def aguardar_inicio_dia():
    now = datetime.now()
    proximo_inicio = datetime.combine(now.date(), datetime.min.time()) + timedelta(hours=HORA_INICIO)
    if now.hour >= HORA_FIM: proximo_inicio += timedelta(days=1)
    espera_segundos = (proximo_inicio - now).total_seconds()
    espera_horas = int(espera_segundos // 3600)
    espera_minutos = int((espera_segundos % 3600) // 60)
    logger.info(f"🌙 Fora do horário comercial. Modo plantão ativado.")
    logger.info(f"⏰ Próxima retomada em {espera_horas}h e {espera_minutos}min.")
    time.sleep(espera_segundos)
    logger.info("☀️ Horário comercial iniciado. Retomando operações.")


def executar_limpeza_semanal():
    logger.warning("🧹 Iniciando limpeza semanal dos perfis do Chrome...")
    for sessao in SESSOES:
        if os.path.exists(sessao["user_data_dir"]):
            try:
                shutil.rmtree(sessao["user_data_dir"])
                logger.info(f"   - Perfil '{sessao['nome']}' limpo com sucesso.")
            except Exception as e:
                logger.error(f"   - Erro ao limpar perfil '{sessao['nome']}': {e}")
    logger.warning("⚠️ Limpeza concluída. Um novo login será necessário na próxima execução.")


# =========================
# TRATAMENTO DE RESPOSTAS E FINALIZAÇÃO (v5.5 com Pausa Adaptativa)
# =========================
def classificar_resposta(txt):
    t = (txt or "").lower()
    sim = {"sim", "tenho interesse", "quero", "interessado", "ok"}
    nao = {"nao", "não", "não quero", "nao quero", "pare", "sem interesse", "n"}
    if any(p in t for p in sim): return "SIM"
    if any(p in t for p in nao): return "NAO"
    return "OUTRO"


def get_msgs(driver):
    try:
        return driver.find_elements(By.XPATH, XPATH_MSG_IN)
    except Exception:
        return []


def tratar_resposta(driver, numero, sessao):
    if queue_manager.state.get(numero, {}).get("status") == "finalized": return False
    try:
        abrir_chat(driver, numero)
        msgs = get_msgs(driver)
        if not msgs: return False
        texto = (msgs[-1].text or "").strip()
        if not texto: return False
        tipo = classificar_resposta(texto)
        logger.info(f"📩 RESPOSTA RECEBIDA de {numero}: '{texto}' ({tipo})")
        ts = datetime.utcnow().isoformat()
        if dentro_do_horario():
            if tipo == "SIM":
                enviar_texto_seguro(driver, f"{MSG_INTERESSADO}\n\n📝 Formulário de cadastro: {LINK_FORM}")
                salvar_print(driver, f"sim_{numero[-4:]}")
                # v5.5: Pausa adaptativa após interação positiva
                pausa_extra = random.uniform(20, 45)
                logger.info(f"💬 Pausa pós-interação positiva ({pausa_extra:.0f}s)")
                time.sleep(pausa_extra)

                if not ja_registrado(numero, "RESPONDEU_SIM"):
                    escrever_log(ts, numero, "RESPONDEU_SIM", "Link do formulário enviado", sessao["nome"])
                    registrar_hash_evento(numero, "RESPONDEU_SIM")
                return True
            elif tipo == "NAO":
                enviar_texto_seguro(driver, MSG_NAO_INTERESSADO)
                queue_manager.add_to_dnc(numero)
                salvar_print(driver, f"nao_{numero[-4:]}")
                if not ja_registrado(numero, "RESPONDEU_NAO"):
                    escrever_log(ts, numero, "RESPONDEU_NAO", "", sessao["nome"])
                    registrar_hash_evento(numero, "RESPONDEU_NAO")
                return True
            else:
                enviar_texto_seguro(driver, MSG_PEDIR_CONF)
                salvar_print(driver, f"confirma_{numero[-4:]}")
                if not ja_registrado(numero, "PEDIU_CONFIRMACAO"):
                    escrever_log(ts, numero, "PEDIU_CONFIRMACAO", "", sessao["nome"])
                    registrar_hash_evento(numero, "PEDIU_CONFIRMACAO")
                return True
        else:
            msg = fora_horario_mensagem()
            enviar_texto_seguro(driver, msg)
            salvar_print(driver, f"fora_horario_{numero[-4:]}")
            if not ja_registrado(numero, "AUTO_REPLY_FORA_HORARIO"):
                escrever_log(ts, numero, "AUTO_REPLY_FORA_HORARIO", "", sessao["nome"])
                registrar_hash_evento(numero, "AUTO_REPLY_FORA_HORARIO")
            return True
    except Exception as e:
        logger.error(f"Erro ao tratar resposta de {numero}: {e}")
        salvar_print(driver, f"erro_resposta_{numero[-4:]}")
        return False


def finalizar_contato(driver, numero, sessao):
    try:
        if queue_manager.state.get(numero, {}).get("status") == "finalized":
            logger.info(f"⏭️  Contato {numero} já foi finalizado.")
            return False
        abrir_chat(driver, numero)
        enviar_texto_seguro(driver, MSG_FORMULARIO_RECEBIDO)
        salvar_print(driver, f"finalizado_{numero[-4:]}")
        ts = datetime.utcnow().isoformat()
        if not ja_registrado(numero, "FORMULARIO_PREENCHIDO"):
            escrever_log(ts, numero, "FORMULARIO_PREENCHIDO", "Cadastro recebido e finalizado", sessao["nome"])
            registrar_hash_evento(numero, "FORMULARIO_PREENCHIDO")
        queue_manager.mark_finalized(numero)
        logger.info(f"✅ Contato {numero} finalizado com sucesso!")
        return True
    except Exception as e:
        logger.error(f"Erro ao finalizar contato {numero}: {e}")
        return False


# =========================
# PROCESSAMENTO DE EMAILS (se habilitado)
# =========================
def conectar_email():
    try:
        mail = imaplib.IMAP4_SSL(EMAIL_IMAP_SERVER, EMAIL_IMAP_PORT)
        mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        return mail
    except Exception as e:
        logger.error(f"❌ Erro ao conectar ao email: {e}")
        return None


def extrair_numero_telefone(corpo_email):
    padroes = [r'\(\d{2}\)\s*\d{4,5}-\d{4}', r'\d{2}\s*\d{4,5}-\d{4}', r'\d{10,11}']
    for padrao in padroes:
        match = re.search(padrao, corpo_email)
        if match:
            numero = re.sub(r'\D', '', match.group())
            if not numero.startswith("55"): numero = "55" + numero
            return "+" + numero
    return None


def processar_emails(drivers, sessoes):
    if not FINALIZAR_COM_FORMULARIO or not EMAIL_ADDRESS or not EMAIL_PASSWORD: return
    mail = conectar_email()
    if not mail: return
    try:
        mail.select(EMAIL_FOLDER)
        status, messages = mail.search(None, f'(UNSEEN SUBJECT "{EMAIL_SUBJECT_FILTER}")')
        if status != "OK" or not messages[0]:
            logger.info("ℹ️ Nenhum novo email de notificação encontrado.")
            return
        email_ids = messages[0].split()
        logger.info(f"📧 Encontrados {len(email_ids)} emails de notificação.")
        novos_processados = 0
        for email_id in email_ids:
            status, msg_data = mail.fetch(email_id, "(RFC822)")
            if status != "OK": continue
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            timestamp = email.utils.parsedate_to_datetime(msg["Date"])
            timestamp_str = timestamp.isoformat()
            hash_email = gerar_hash_local(email_id.decode(), "EMAIL_PROCESSADO", timestamp_str)
            if not hash_email or hash_email in processed_hashes: continue
            processed_hashes.add(hash_email)
            salvar_hashes_processados()
            corpo = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        corpo = part.get_payload(decode=True).decode()
                        break
            else:
                corpo = msg.get_payload(decode=True).decode()
            numero = extrair_numero_telefone(corpo)
            if not numero:
                logger.warning(f"⚠️ Não foi possível extrair o número de telefone do email {email_id.decode()}")
                continue
            logger.info(f"🔔 Novo formulário detectado no email para {numero}. Finalizando contato...")
            for driver in drivers:
                if finalizar_contato(driver, numero, sessoes[0]):
                    novos_processados += 1
                    break
            mail.store(email_id, '+FLAGS', '\\Seen')
        if novos_processados > 0: logger.info(f"✅ {novos_processados} novos formulários processados.")
    except Exception as e:
        logger.error(f"❌ Erro ao processar emails: {e}")
    finally:
        try:
            mail.logout()
        except:
            pass


# =========================
# LOOP PRINCIPAL (v5.5)
# =========================
def main():
    # v5.5: Verificação de dependências e exibição da tabela inicial
    checar_dependencias()
    exibir_tabela_inicial()

    logger.info("🚀 Iniciando WhatsApp_v5.5_defensive")
    logger.info(
        "📅 Recursos: Fila persistente, driver seguro, logs avançados, performance (Polars), monitoramento (Rich)")

    queue_manager.init_state_from_sheet()
    carregar_hashes_processados()

    if LIMPEZA_SEMANAL_ATIVA and datetime.now().weekday() == DIA_LIMPEZA_SEMANAL and datetime.now().hour == HORA_LIMPEZA_SEMANAL:
        executar_limpeza_semanal()
        logger.info("🔄 Reiniciando após limpeza semanal...")
        time.sleep(5)
        os.execv(sys.executable, ['python'] + sys.argv)

    SESSOES_ATIVAS = [SESSOES[0]]
    LIMITE_POR_NUMERO = 175
    if len(queue_manager.state) > LIMITE_POR_NUMERO: SESSOES_ATIVAS = SESSOES

    drivers = []
    for sessao in SESSOES_ATIVAS:
        criar_perfil_se_nao_existir(sessao)
        driver = criar_driver_discreto(sessao)
        abrir_whatsapp(driver)
        drivers.append(driver)
        time.sleep(3)

    dia_atual = date.today()
    ultimo_evento = datetime.now()
    ultima_checagem_email = time.time()
    contador_envios = 0
    ultima_limpeza_prints = time.time()

    try:
        while True:
            # v5.5: Limpeza periódica de prints
            if time.time() - ultima_limpeza_prints > 86400:  # A cada 24h
                limpar_prints_antigos()
                ultima_limpeza_prints = time.time()

            if REINICIAR_DIARIAMENTE and date.today() != dia_atual:
                logger.info("🌅 Novo dia detectado — encerrando e reiniciando o bot.")
                for d in drivers: d.quit()
                rel = gerar_relatorio_diario()
                logger.info(f"📊 Relatório do dia {dia_atual} finalizado: {json.dumps(rel, indent=2)}")
                logger.info("🔄 Reiniciando script para o novo dia...")
                time.sleep(5)
                os.execv(sys.executable, ['python'] + sys.argv)

            if MODO_PLANTAO and not dentro_do_horario():
                for d in drivers: d.quit()
                aguardar_inicio_dia()
                drivers = []
                for sessao in SESSOES_ATIVAS:
                    criar_perfil_se_nao_existir(sessao)
                    driver = criar_driver_discreto(sessao)
                    abrir_whatsapp(driver)
                    drivers.append(driver)
                    time.sleep(3)
                continue

            if FINALIZAR_COM_FORMULARIO and (time.time() - ultima_checagem_email > INTERVALO_CHECAGEM_EMAIL_SEG):
                processar_emails(drivers, SESSOES_ATIVAS)
                ultima_checagem_email = time.time()

            ativo = False
            for i, driver in enumerate(drivers):
                sessao = SESSOES_ATIVAS[i]

                # v5.5: Verifica se o processo Chrome está ativo
                if not chrome_ativo():
                    logger.warning("🚨 Processo Chrome não encontrado. Reiniciando driver...")
                    driver.quit()
                    driver = criar_driver_discreto(sessao)
                    abrir_whatsapp(driver)
                    drivers[i] = driver
                    time.sleep(5)
                    continue

                if not sessao_ativa(driver):
                    logger.warning("⚠️ Sessão desconectada. Reiniciando driver...")
                    driver.quit()
                    driver = criar_driver_discreto(sessao)
                    abrir_whatsapp(driver)
                    drivers[i] = driver
                    time.sleep(5)
                    continue

                numero, meta = queue_manager.get_next_number_for_session(sessao["nome"])
                if not numero:
                    logger.info(f"Nenhum número disponível para a sessão {sessao['nome']}")
                    continue

                try:
                    abrir_chat(driver, numero)
                    mensagem = escolher_mensagem()
                    enviar_texto_seguro(driver, mensagem)
                    ts = datetime.utcnow().isoformat()
                    if not ja_registrado(numero, "ENVIO_OK"):
                        escrever_log(ts, numero, "ENVIO_OK", mensagem, sessao["nome"])
                        registrar_hash_evento(numero, "ENVIO_OK")
                    queue_manager.mark_sent(numero, sessao["nome"])
                    salvar_print(driver, f"envio_{numero[-4:]}")
                    agendar_reenvio(numero, horas=12)
                    simular_acao_humana(driver)
                    contador_envios += 1
                    ativo = True
                    ultimo_evento = datetime.now()

                    pausa = escolher_pausa()
                    logger.info(f"⏳ Pausando por {pausa:.1f}s antes do próximo envio...")
                    time.sleep(pausa)

                    if contador_envios > 0 and contador_envios % 150 == 0:
                        logger.info("🔁 Reiniciando driver automaticamente para evitar travamentos...")
                        driver.quit()
                        time.sleep(10)
                        driver = criar_driver_discreto(sessao)
                        abrir_whatsapp(driver)
                        drivers[i] = driver

                    if contador_envios % random.randint(30, 50) == 0:
                        pausa_longa = random.uniform(900, 1500)
                        logger.info(f"😴 Pausa longa simulada de {pausa_longa / 60:.1f} minutos.")
                        time.sleep(pausa_longa)

                except Exception as e:
                    logger.error(f"Erro ao enviar para {numero}: {e}")
                    queue_manager.mark_failed(numero, sessao["nome"], reason=str(e))
                    salvar_print(driver, f"erro_envio_{numero[-4:]}")
                    continue

                numeros_para_responder = [n for n in queue_manager.state.keys() if
                                          queue_manager.state[n].get("status") == "sent"]
                for n in numeros_para_responder:
                    if tratar_resposta(driver, n, sessao):
                        ativo = True
                        ultimo_evento = datetime.now()

                processar_agenda_reenvio(driver)
                time.sleep(5)

            if not ativo and (datetime.now() - ultimo_evento).seconds > INATIVIDADE_AUTOEXIT_MIN * 60:
                logger.info("🕒 Inatividade detectada — encerrando com segurança.")
                break

    except KeyboardInterrupt:
        logger.warning("🛑 Interrompido manualmente (Ctrl+C).")
    except Exception as e:
        logger.error(f"❌ Erro crítico: {e}")
    finally:
        for d in drivers:
            try:
                d.quit()
            except:
                pass
        rel = gerar_relatorio_diario()
        logger.info(f"📊 Relatório final: {json.dumps(rel, indent=2)}")
        logger.info("🏁 Bot finalizado.")


if __name__ == "__main__":
    main()