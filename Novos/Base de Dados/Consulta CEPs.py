# -*- coding: utf-8 -*-
"""
======================================================
📦 Consulta Automática de CEPs via API ViaCEP
Versão: 3.0 (2025-10-16)
Autor: bb-assistente 😎
------------------------------------------------------
✅ Localiza o arquivo Excel automaticamente
✅ Normaliza e valida CEPs
✅ Consulta API ViaCEP com cache + tentativas
✅ Gera planilha formatada + resumo por UF
✅ Cria log detalhado com tempo total e erros
======================================================
"""

import os
import time
import random
import logging
import pandas as pd
import requests
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ======================================================
# ⚙️ CONFIGURAÇÕES
# ======================================================

PASTA_ENTRADA = r"C:\Users\J&T-099\OneDrive - Speed Rabbit Express Ltda (1)\Área de Trabalho\Testes\Local de Teste\CEP"
ARQUIVO_SAIDA = os.path.join(PASTA_ENTRADA, "CEPs_Resultados.xlsx")
ARQUIVO_LOG = os.path.join(PASTA_ENTRADA, f"log_{time.strftime('%Y%m%d')}.txt")

# Configuração de logs
logging.basicConfig(
    filename=ARQUIVO_LOG,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8"
)

# ======================================================
# 🔍 FUNÇÃO PARA LOCALIZAR O ARQUIVO
# ======================================================

def encontrar_excel(pasta: str) -> str | None:
    """Procura automaticamente o primeiro arquivo .xlsx na pasta."""
    print(f"🔎 Procurando arquivo Excel em:\n{pasta}\n")

    if not os.path.exists(pasta):
        print("❌ A pasta informada não existe.")
        return None

    arquivos = [f for f in os.listdir(pasta) if f.lower().endswith(".xlsx")]
    if not arquivos:
        print("⚠️ Nenhum arquivo .xlsx encontrado.")
        return None

    if len(arquivos) > 1:
        print("📁 Vários arquivos encontrados:")
        for i, nome in enumerate(arquivos, 1):
            print(f"  {i}. {nome}")
        try:
            escolha = int(input("\nDigite o número do arquivo desejado: "))
            caminho = os.path.join(pasta, arquivos[escolha - 1])
        except (ValueError, IndexError):
            print("⚠️ Escolha inválida. Usando o primeiro arquivo.")
            caminho = os.path.join(pasta, arquivos[0])
    else:
        caminho = os.path.join(pasta, arquivos[0])

    print(f"✅ Arquivo selecionado: {os.path.basename(caminho)}\n")
    return caminho

# ======================================================
# 🌐 CONSULTA API VIACEP
# ======================================================

def consulta_viacep(cep: str):
    """Consulta cidade e UF usando a API ViaCEP (com retry e backoff)."""
    url = f"https://viacep.com.br/ws/{cep}/json/"
    for tentativa in range(3):
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            data = r.json()
            if data.get("erro"):
                return None, None, "CEP não encontrado"
            return data.get("localidade"), data.get("uf"), "OK"
        except Exception as e:
            time.sleep(random.uniform(0.6, 1.2))
    return None, None, "Erro permanente"

# ======================================================
# 🎨 FORMATAÇÃO DO EXCEL
# ======================================================

def formatar_excel(caminho: str):
    """Aplica formatação visual no Excel."""
    wb = load_workbook(caminho)
    ws = wb["CEPs Detalhados"]

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    negrito = Font(bold=True)

    for cell in ws["D"][1:]:
        if cell.value == "OK":
            for c in ws[cell.row]:
                c.fill = verde
        elif cell.value != "OK":
            for c in ws[cell.row]:
                c.fill = vermelho

    for cell in ws[1]:
        cell.font = negrito

    wb.save(caminho)

# ======================================================
# 🧠 FUNÇÃO PRINCIPAL
# ======================================================

def main():
    inicio = time.time()
    caminho = encontrar_excel(PASTA_ENTRADA)
    if not caminho:
        return

    print("📂 Lendo planilha...\n")
    df = pd.read_excel(caminho, dtype={"CEP": str})

    if "CEP" not in df.columns:
        print("❌ Erro: a planilha não contém uma coluna chamada 'CEP'.")
        return

    # Normaliza os CEPs
    df["CEP"] = df["CEP"].astype(str).str.replace(r'\D', '', regex=True).str.zfill(8)

    cidades, ufs, status = [], [], []
    cache_cep = {}

    print(f"🔎 Consultando {len(df)} CEPs no ViaCEP...\n")
    for cep in tqdm(df["CEP"], desc="Consultando CEPs", ncols=80):
        if not cep.isdigit() or len(cep) != 8:
            cidades.append("")
            ufs.append("")
            status.append("CEP inválido")
            continue

        if cep in cache_cep:
            cidade, uf, st = cache_cep[cep]
        else:
            cidade, uf, st = consulta_viacep(cep)
            cache_cep[cep] = (cidade, uf, st)

        cidades.append(cidade or "")
        ufs.append(uf or "")
        status.append(st)

    resultado = pd.DataFrame({
        "CEP": df["CEP"],
        "Cidade": cidades,
        "UF": ufs,
        "Status da consulta": status
    })

    # Cria resumo por UF
    resumo = resultado.groupby("UF", dropna=True).size().reset_index(name="Qtd_CEPs")

    with pd.ExcelWriter(ARQUIVO_SAIDA, engine='openpyxl') as writer:
        resultado.to_excel(writer, index=False, sheet_name="CEPs Detalhados")
        resumo.to_excel(writer, index=False, sheet_name="Resumo por UF")

    formatar_excel(ARQUIVO_SAIDA)

    tempo_total = round(time.time() - inicio, 2)
    total_erros = (resultado["Status da consulta"] != "OK").sum()

    print(f"\n✅ Planilha gerada com sucesso: {ARQUIVO_SAIDA}")
    print(f"📊 CEPs únicos: {resultado['CEP'].nunique()} / Totais: {len(resultado)}")
    print(f"❌ CEPs inválidos/erro: {total_erros}")
    print(f"⏱️ Tempo total: {tempo_total}s\n")

    logging.info(f"Arquivo processado: {os.path.basename(caminho)}")
    logging.info(f"Total de CEPs: {len(resultado)} | Erros: {total_erros}")
    logging.info(f"Tempo total: {tempo_total}s")


# ======================================================
# 🚀 EXECUÇÃO
# ======================================================
if __name__ == "__main__":
    main()
