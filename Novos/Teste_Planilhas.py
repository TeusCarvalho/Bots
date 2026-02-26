from __future__ import annotations

import sys
import csv
from pathlib import Path
from datetime import datetime

# ======== CONFIG (ajuste aqui se quiser) ========
PASTA_ALVO = r"C:\Users\J&T-099\OneDrive - Speed Rabbit Express Ltda\QUALIDADE_ FILIAL GO - BASE DE DADOS\01. OPERAÇÃO CD\11. ENTREGA REALIZADA - LISTA"
RECURSIVO = True  # True = entra em subpastas também
EXTENSOES = {".xlsx", ".xlsm", ".xlsb", ".xls"}

# Saída (relatório)
GERAR_CSV = True
NOME_CSV = "relatorio_abas_excel.csv"
# ===============================================


def list_sheets_xlsx_xlsm(path: Path) -> list[str]:
    # .xlsx / .xlsm via openpyxl (mais leve pra só pegar sheetnames)
    from openpyxl import load_workbook  # pip install openpyxl
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def list_sheets_xlsb(path: Path) -> list[str]:
    # .xlsb via pandas + pyxlsb
    import pandas as pd  # pip install pandas
    xf = pd.ExcelFile(path, engine="pyxlsb")  # pip install pyxlsb
    return list(xf.sheet_names)


def list_sheets_xls(path: Path) -> list[str]:
    # .xls antigo via pandas + xlrd
    import pandas as pd  # pip install pandas
    xf = pd.ExcelFile(path, engine="xlrd")  # pip install xlrd
    return list(xf.sheet_names)


def get_sheetnames(file_path: Path) -> tuple[list[str] | None, str | None]:
    """
    Retorna (sheetnames, erro)
    - sheetnames: lista de abas se ok
    - erro: mensagem se falhar
    """
    ext = file_path.suffix.lower()

    try:
        if ext in (".xlsx", ".xlsm"):
            return list_sheets_xlsx_xlsm(file_path), None
        if ext == ".xlsb":
            return list_sheets_xlsb(file_path), None
        if ext == ".xls":
            return list_sheets_xls(file_path), None
        return None, f"Extensão não suportada: {ext}"
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"


def iter_excel_files(root: Path):
    if RECURSIVO:
        for p in root.rglob("*"):
            if p.is_file() and p.suffix.lower() in EXTENSOES:
                yield p
    else:
        for p in root.glob("*"):
            if p.is_file() and p.suffix.lower() in EXTENSOES:
                yield p


def main() -> int:
    root = Path(PASTA_ALVO)

    if not root.exists():
        print(f"❌ Pasta não existe:\n{root}")
        return 1

    files = list(iter_excel_files(root))
    if not files:
        print(f"⚠️ Não encontrei arquivos Excel em:\n{root}")
        return 0

    print(f"📁 Pasta alvo: {root}")
    print(f"📄 Arquivos Excel encontrados: {len(files)}\n")

    rows = []
    ok = 0
    fail = 0

    for i, f in enumerate(files, start=1):
        sheetnames, err = get_sheetnames(f)

        if sheetnames is not None:
            ok += 1
            abas = " | ".join(sheetnames)
            print(f"[{i}/{len(files)}] ✅ {f.name}")
            print(f"    Abas ({len(sheetnames)}): {abas}\n")
            rows.append({
                "arquivo": str(f),
                "nome_arquivo": f.name,
                "extensao": f.suffix.lower(),
                "qtd_abas": len(sheetnames),
                "abas": abas,
                "erro": ""
            })
        else:
            fail += 1
            print(f"[{i}/{len(files)}] ❌ {f.name}")
            print(f"    ERRO: {err}\n")
            rows.append({
                "arquivo": str(f),
                "nome_arquivo": f.name,
                "extensao": f.suffix.lower(),
                "qtd_abas": "",
                "abas": "",
                "erro": err or ""
            })

    print("======== RESUMO ========")
    print(f"✅ OK:   {ok}")
    print(f"❌ Erro: {fail}")
    print("========================\n")

    if GERAR_CSV:
        out_csv = Path.cwd() / NOME_CSV
        with out_csv.open("w", newline="", encoding="utf-8") as fp:
            writer = csv.DictWriter(fp, fieldnames=["arquivo", "nome_arquivo", "extensao", "qtd_abas", "abas", "erro"])
            writer.writeheader()
            writer.writerows(rows)

        print(f"📌 CSV gerado: {out_csv}")

    return 0


if __name__ == "__main__":
    sys.exit(main())