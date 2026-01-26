# -*- coding: utf-8 -*-
"""
Comparativo consolidado em UM único Excel, pensado para rodar via "Run" no PyCharm
(sem PowerShell / sem parâmetros).

O script procura os arquivos Excel na pasta fixa (BASE_DIR) e gera:
- comparativo_unico.xlsx

Saídas no Excel:
- Resumo        : por Transportadora x Aba (quantidade comparada, quem foi menor, etc.)
- Comparativo   : detalhe de todas as chaves casadas (valor_nova vs valor_gp)
- NaoCasados    : itens presentes só em uma das fontes

Obs:
- "NOVA" = arquivo da tabela matriz (transportadora/peso/destino)
- "GP"   = arquivo com faixas de peso (de/até) por destino
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import openpyxl


# =========================
# CONFIG (para usar só RUN)
# =========================
BASE_DIR = Path(r"C:\Users\J&T-099\OneDrive - Speed Rabbit Express Ltda (1)\Área de Trabalho\Pastas")

# Nomes "preferidos" (se existirem exatamente, usa eles; senão tenta achar por padrão)
ARQ_NOVA_PREFERRED = "Nova Tabela melhor envio J&T Express.finalizada1.xlsx"
ARQ_GP_PREFERRED = "Tabelas GP - C2C 11.2025. filtrado1.xlsx"

ARQ_SAIDA = "comparativo_unico.xlsx"

# Transportadoras que serão comparadas (linhas do arquivo NOVA)
CARRIERS = ["J&T Express Standard", "Menor valor"]

# Tolerância para considerar "igual"
TOL_IGUAL = 0.01


# =========================
# Helpers
# =========================
def norm_sheet_name(name: str) -> str:
    s = str(name).upper().strip()
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def to_float(x) -> float:
    """Converte valores como 'R$ 35,64' / '35.64' / numérico -> float."""
    if x is None:
        return np.nan

    if isinstance(x, (int, float, np.integer, np.floating)):
        try:
            if isinstance(x, float) and math.isnan(x):
                return np.nan
        except Exception:
            pass
        return float(x)

    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return np.nan

    # NBSP
    s = s.replace("\u00A0", " ")
    # moeda
    s = s.replace("R$", "").replace("r$", "").strip()

    # pt-BR (milhar e decimal)
    if "," in s and "." in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    # mantém dígitos/ponto/sinal
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", "-", "."}:
        return np.nan

    try:
        return float(s)
    except Exception:
        return np.nan


def find_cell(df: pd.DataFrame, predicate) -> Optional[Tuple[int, int]]:
    """Retorna (row, col) do primeiro match textual no df (header=None)."""
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            v = df.iat[i, j]
            if isinstance(v, str) and predicate(v):
                return i, j
    return None


def find_excel_file(base_dir: Path, preferred_name: str, name_patterns: List[str]) -> Path:
    """
    Encontra um arquivo Excel:
    1) Se existir o 'preferred_name' exato, usa ele.
    2) Senão, procura por padrões (regex simples) dentro do nome.
    """
    p = base_dir / preferred_name
    if p.exists():
        return p

    candidates = sorted(base_dir.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"Nenhum .xlsx encontrado em: {base_dir}")

    for pat in name_patterns:
        rx = re.compile(pat, flags=re.IGNORECASE)
        for c in candidates:
            if rx.search(c.name):
                return c

    msg = "Não encontrei o arquivo esperado.\n"
    msg += f"Pasta: {base_dir}\n"
    msg += "Arquivos .xlsx encontrados:\n"
    for c in candidates[:50]:
        msg += f" - {c.name}\n"
    msg += "\nDica: renomeie o arquivo para o nome esperado ou ajuste os padrões no código."
    raise FileNotFoundError(msg)


# =========================
# Parsers
# =========================
def parse_nova_tabela(path: Path, sheet: str, carriers: List[str]) -> pd.DataFrame:
    """
    Parse da planilha NOVA (matriz):
    - Localiza 'ESTADO' (linha de destinos)
    - Localiza 'PESO' (coluna do peso)
    - Transportadora costuma estar na coluna 0
    - Valores por destino nas colunas à direita
    """
    df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")

    loc_estado = find_cell(df, lambda s: s.strip().upper() == "ESTADO")
    loc_peso = find_cell(df, lambda s: s.strip().upper() == "PESO")
    if not loc_estado or not loc_peso:
        return pd.DataFrame(columns=["aba", "transportadora", "peso_ate", "destino", "valor"])

    row_estado, col_estado = loc_estado
    row_peso, col_peso = loc_peso

    col_start = max(col_estado, col_peso) + 1

    dests: Dict[int, str] = {}
    for col in range(col_start, df.shape[1]):
        v = df.iat[row_estado, col]
        if isinstance(v, str) and v.strip():
            dests[col] = v.strip().upper()

    carrier_map = {c.strip().upper(): c for c in carriers}

    records = []
    for i in range(row_peso + 1, df.shape[0]):
        carrier = df.iat[i, 0]
        if not isinstance(carrier, str):
            continue

        c_up = carrier.strip().upper()
        if c_up not in carrier_map:
            continue

        peso_f = to_float(df.iat[i, col_peso])
        if math.isnan(peso_f):
            continue

        for col, destino in dests.items():
            val = to_float(df.iat[i, col])
            if math.isnan(val):
                continue
            records.append((norm_sheet_name(sheet), carrier_map[c_up], float(peso_f), destino, val))

    out = pd.DataFrame(records, columns=["aba", "transportadora", "peso_ate", "destino", "valor"])

    # Dedup defensivo: se repetir a mesma chave, mantém o menor valor
    if not out.empty:
        out = (
            out.groupby(["aba", "transportadora", "destino", "peso_ate"], as_index=False)
            .agg(valor=("valor", "min"))
        )

    return out


def parse_tabela_gp(path: Path, sheet: str) -> pd.DataFrame:
    """
    Parse da planilha GP:
    - Procura 'Faixa de Peso em Kg'
    - Procura 'GEOCOM' acima (header de destinos)
    - Lê (peso_de, peso_ate) e valores por destino
    """
    df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")

    faixa_rows: List[int] = []
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            v = df.iat[i, j]
            if isinstance(v, str):
                s = v.strip().upper()
                if "FAIXA" in s and "PESO" in s:
                    faixa_rows.append(i)
                    break

    records = []
    for row_faixa in faixa_rows:
        dest_row = None
        for k in range(row_faixa - 1, max(-1, row_faixa - 15), -1):
            row_vals = df.iloc[k, :].tolist()
            if any(isinstance(x, str) and x.strip().upper() == "GEOCOM" for x in row_vals):
                dest_row = k
                break
        if dest_row is None:
            dest_row = row_faixa - 1

        dests: Dict[int, str] = {}
        for col in range(3, df.shape[1]):
            v = df.iat[dest_row, col]
            if isinstance(v, str) and v.strip():
                dests[col] = v.strip().upper()

        i = row_faixa + 1
        while i < df.shape[0]:
            lo_raw = df.iat[i, 1]
            hi_raw = df.iat[i, 2]

            lo_f = to_float(lo_raw)
            hi_f = to_float(hi_raw)

            # condição de parada: linha vazia (sem faixa e sem valores)
            if (lo_raw is None) or (isinstance(lo_raw, float) and math.isnan(lo_raw)) or (
                isinstance(lo_raw, str) and not lo_raw.strip()
            ):
                row_vals = df.iloc[i, 3:3 + len(dests)].tolist() if dests else df.iloc[i, :].tolist()
                if all(
                    (x is None)
                    or (isinstance(x, float) and math.isnan(x))
                    or (isinstance(x, str) and not x.strip())
                    for x in row_vals
                ):
                    break

            if not math.isnan(hi_f):
                for col, destino in dests.items():
                    val = to_float(df.iat[i, col])
                    if math.isnan(val):
                        continue
                    records.append(
                        (
                            norm_sheet_name(sheet),
                            float(lo_f) if not math.isnan(lo_f) else np.nan,
                            float(hi_f),
                            destino,
                            val,
                        )
                    )

            i += 1

    out = pd.DataFrame(records, columns=["aba", "peso_de", "peso_ate", "destino", "valor"])

    # Dedup defensivo: mesma chave -> mantém o menor valor
    if not out.empty:
        out = (
            out.groupby(["aba", "destino", "peso_ate"], as_index=False)
            .agg(valor=("valor", "min"), peso_de=("peso_de", "min"))
        )

    return out


# =========================
# Comparação + Saída
# =========================
def compare_one_carrier(nova_long: pd.DataFrame, gp_long: pd.DataFrame, carrier_name: str) -> pd.DataFrame:
    """
    Compara uma transportadora específica do NOVA contra GP.
    Chave: aba + destino + peso_ate

    Correção aplicada:
    - Se 'transportadora' já existir após merge, sobrescreve em vez de inserir.
    """
    a = nova_long[nova_long["transportadora"] == carrier_name].copy()
    b = gp_long.copy()

    m = a.merge(
        b,
        on=["aba", "destino", "peso_ate"],
        how="outer",
        suffixes=("_nova", "_gp"),
        indicator=True,
    )

    # ---- CORREÇÃO: não inserir se já existir; apenas garantir o valor ----
    if "transportadora" in m.columns:
        m["transportadora"] = carrier_name
    else:
        m.insert(0, "transportadora", carrier_name)

    m["diff"] = m["valor_nova"] - m["valor_gp"]

    def decide(row):
        if pd.isna(row.get("valor_nova")) or pd.isna(row.get("valor_gp")):
            return "NAO_CASADO"
        d = float(row["diff"])
        if abs(d) <= TOL_IGUAL:
            return "IGUAL"
        return "NOVA" if d < 0 else "GP"

    m["menor"] = m.apply(decide, axis=1)
    return m


def build_resumo(comp_all: pd.DataFrame) -> pd.DataFrame:
    both = comp_all[comp_all["_merge"] == "both"].copy()
    if both.empty:
        return pd.DataFrame(
            columns=["transportadora", "aba", "qtd_celulas", "nova_menor", "gp_menor", "iguais", "diff_media", "diff_max_abs"]
        )

    # cria flags para somar sem apply
    both["is_nova_menor"] = (both["menor"] == "NOVA").astype(int)
    both["is_gp_menor"] = (both["menor"] == "GP").astype(int)
    both["is_igual"] = (both["menor"] == "IGUAL").astype(int)
    both["abs_diff"] = both["diff"].abs()

    res = (
        both.groupby(["transportadora", "aba"], as_index=False)
        .agg(
            qtd_celulas=("diff", "size"),
            nova_menor=("is_nova_menor", "sum"),
            gp_menor=("is_gp_menor", "sum"),
            iguais=("is_igual", "sum"),
            diff_media=("diff", "mean"),
            diff_max_abs=("abs_diff", "max"),
        )
        .sort_values(["transportadora", "diff_max_abs", "qtd_celulas"], ascending=[True, False, False])
    )

    return res


def build_comparativo(comp_all: pd.DataFrame) -> pd.DataFrame:
    both = comp_all[comp_all["_merge"] == "both"].copy()
    both["abs_diff"] = both["diff"].abs()
    cols = ["transportadora", "aba", "destino", "peso_ate", "valor_nova", "valor_gp", "diff", "menor", "abs_diff"]
    return both[cols].sort_values(["transportadora", "abs_diff", "aba", "destino", "peso_ate"], ascending=[True, False, True, True, True])


def build_nao_casados(comp_all: pd.DataFrame) -> pd.DataFrame:
    left = comp_all[comp_all["_merge"] == "left_only"].copy()
    right = comp_all[comp_all["_merge"] == "right_only"].copy()

    left_u = left[["transportadora", "aba", "destino", "peso_ate", "valor_nova"]].rename(columns={"valor_nova": "valor"})
    right_u = right[["transportadora", "aba", "destino", "peso_ate", "valor_gp"]].rename(columns={"valor_gp": "valor"})

    left_u["origem"] = "NOVA"
    right_u["origem"] = "GP"

    u = pd.concat([left_u, right_u], ignore_index=True)
    return u.sort_values(["transportadora", "origem", "aba", "destino", "peso_ate"])


def main():
    if not BASE_DIR.exists():
        raise FileNotFoundError(
            f"BASE_DIR não existe:\n{BASE_DIR}\n\nAjuste a variável BASE_DIR no topo do script."
        )

    nova_path = find_excel_file(
        BASE_DIR,
        ARQ_NOVA_PREFERRED,
        name_patterns=[
            r"^Nova\s+Tabela\s+melhor\s+envio.*\.xlsx$",
            r"Nova\s+Tabela\s+melhor\s+envio",
        ],
    )

    gp_path = find_excel_file(
        BASE_DIR,
        ARQ_GP_PREFERRED,
        name_patterns=[
            r"^Tabelas\s+GP\s*-\s*C2C.*\.xlsx$",
            r"Tabelas\s+GP",
            r"C2C",
        ],
    )

    out_path = BASE_DIR / ARQ_SAIDA

    print("==============================================")
    print("[OK] Base:", BASE_DIR)
    print("[OK] NOVA:", nova_path.name)
    print("[OK] GP  :", gp_path.name)
    print("[OK] OUT :", out_path.name)
    print("[OK] Carriers:", ", ".join(CARRIERS))
    print("==============================================")

    # Parse NOVA
    wb_nova = openpyxl.load_workbook(nova_path, read_only=True, data_only=True)
    nova_parts = []
    for sh in wb_nova.sheetnames:
        part = parse_nova_tabela(nova_path, sh, CARRIERS)
        if not part.empty:
            nova_parts.append(part)
    nova_long = pd.concat(nova_parts, ignore_index=True) if nova_parts else pd.DataFrame()

    # Parse GP
    wb_gp = openpyxl.load_workbook(gp_path, read_only=True, data_only=True)
    gp_parts = []
    for sh in wb_gp.sheetnames:
        part = parse_tabela_gp(gp_path, sh)
        if not part.empty:
            gp_parts.append(part)
    gp_long = pd.concat(gp_parts, ignore_index=True) if gp_parts else pd.DataFrame()

    if nova_long.empty:
        raise RuntimeError("Não consegui extrair dados do arquivo NOVA (não achei 'ESTADO' e/ou 'PESO' nas abas).")
    if gp_long.empty:
        raise RuntimeError("Não consegui extrair dados do arquivo GP (não achei 'Faixa de Peso em Kg' nas abas).")

    # Comparação por transportadora
    comps = []
    for carrier in CARRIERS:
        comps.append(compare_one_carrier(nova_long, gp_long, carrier))
    comp_all = pd.concat(comps, ignore_index=True) if comps else pd.DataFrame()

    resumo = build_resumo(comp_all)
    comparativo = build_comparativo(comp_all)
    nao_casados = build_nao_casados(comp_all)

    # Export
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        resumo.to_excel(writer, index=False, sheet_name="Resumo")
        comparativo.to_excel(writer, index=False, sheet_name="Comparativo")
        nao_casados.to_excel(writer, index=False, sheet_name="NaoCasados")

    print("==============================================")
    print(f"[OK] Gerado: {out_path}")
    print("==============================================")


if __name__ == "__main__":
    main()
